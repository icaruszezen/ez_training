"""标注指导文档导出页面。

支持多选数据集和标签类型，自动抽样截取目标区域（扩大范围），
最终导出包含指导图像的 Excel 文件。
"""

import io
import logging
import os
import random
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Set

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from PyQt5.QtCore import QThread, Qt, pyqtSignal
from PyQt5.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QListWidget,
    QListWidgetItem,
    QVBoxLayout,
    QWidget,
)
from qfluentwidgets import (
    BodyLabel,
    CardWidget,
    CaptionLabel,
    DoubleSpinBox,
    FluentIcon as FIF,
    InfoBar,
    InfoBarPosition,
    PrimaryPushButton,
    ProgressBar,
    PushButton,
    ScrollArea,
    SpinBox,
    SubtitleLabel,
    TitleLabel,
)

from ez_training.common.constants import SUPPORTED_IMAGE_FORMATS
from ez_training.common.image_utils import imread_unicode

logger = logging.getLogger(__name__)

_SHEET_NAME_INVALID_RE = re.compile(r'[/\\*?\[\]:\'"!]')
_IMG_CACHE_MAX = 8


# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------

@dataclass
class BBoxRecord:
    """一条标注框记录，关联图片路径和绝对像素坐标。"""

    image_path: str
    label: str
    x_min: int
    y_min: int
    x_max: int
    y_max: int


# ---------------------------------------------------------------------------
# LabelScanWorker
# ---------------------------------------------------------------------------

class LabelScanWorker(QThread):
    """扫描选中数据集目录，解析 VOC XML / YOLO TXT，收集 label -> [BBoxRecord]。"""

    progress = pyqtSignal(int, int)
    finished = pyqtSignal(dict)  # Dict[str, List[BBoxRecord]]

    def __init__(self, directories: List[str]):
        super().__init__()
        self._directories = list(directories)
        self._cancelled = False

    def cancel(self):
        self._cancelled = True

    # -- helpers --

    @staticmethod
    def _load_classes_for_dir(directory: str) -> List[str]:
        d = Path(directory)
        for candidate in [
            d / "classes.txt",
            d / "labels" / "classes.txt",
            d / ".." / "classes.txt",
        ]:
            if candidate.exists():
                try:
                    with open(candidate, "r", encoding="utf-8") as f:
                        names = [line.strip() for line in f if line.strip()]
                    if names:
                        return names
                except Exception:
                    pass
        return []

    @staticmethod
    def _find_image_by_stem(annotation_path: Path) -> Optional[str]:
        """按标注文件的 stem 查找对应图片，支持 labels/->images/ 和 Annotations/->JPEGImages/ 回退。"""
        stem = annotation_path.stem
        parent = annotation_path.parent

        for ext in SUPPORTED_IMAGE_FORMATS:
            candidate = parent / f"{stem}{ext}"
            if candidate.exists():
                return str(candidate)

        parent_name = parent.name.lower()
        fallback_dirs: List[Path] = []
        if parent_name == "labels":
            fallback_dirs.append(parent.parent / "images")
        elif parent_name == "annotations":
            fallback_dirs.append(parent.parent / "JPEGImages")

        for fb_dir in fallback_dirs:
            if fb_dir.is_dir():
                for ext in SUPPORTED_IMAGE_FORMATS:
                    candidate = fb_dir / f"{stem}{ext}"
                    if candidate.exists():
                        return str(candidate)
        return None

    def _parse_voc(self, xml_path: Path) -> List[BBoxRecord]:
        records: List[BBoxRecord] = []
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
            filename_el = root.find("filename")
            if filename_el is None or not filename_el.text:
                return records
            img_path = str(xml_path.parent / filename_el.text.strip())
            if not os.path.isfile(img_path):
                found = self._find_image_by_stem(xml_path)
                if found is None:
                    return records
                img_path = found
            for obj in root.findall("object"):
                name = (obj.findtext("name") or "").strip()
                bnd = obj.find("bndbox")
                if not name or bnd is None:
                    continue
                try:
                    x_min = int(float((bnd.findtext("xmin") or "0").strip()))
                    y_min = int(float((bnd.findtext("ymin") or "0").strip()))
                    x_max = int(float((bnd.findtext("xmax") or "0").strip()))
                    y_max = int(float((bnd.findtext("ymax") or "0").strip()))
                except ValueError:
                    continue
                records.append(BBoxRecord(
                    image_path=img_path, label=name,
                    x_min=x_min, y_min=y_min, x_max=x_max, y_max=y_max,
                ))
        except Exception:
            logger.debug("Failed to parse VOC: %s", xml_path, exc_info=True)
        return records

    def _parse_yolo(
        self,
        txt_path: Path,
        class_names: List[str],
        size_cache: Dict[str, tuple],
    ) -> List[BBoxRecord]:
        from ez_training.common.annotation_utils import read_yolo_boxes

        records: List[BBoxRecord] = []
        img_path = self._find_image_by_stem(txt_path)
        if img_path is None:
            return records

        cached = size_cache.get(img_path)
        if cached is not None:
            img_w, img_h = cached
        else:
            try:
                with PILImage.open(img_path) as pil_img:
                    img_w, img_h = pil_img.size
                size_cache[img_path] = (img_w, img_h)
            except Exception:
                return records

        for b in read_yolo_boxes(txt_path, img_w, img_h, class_names):
            records.append(BBoxRecord(
                image_path=img_path,
                label=b["label"],
                x_min=b["xmin"],
                y_min=b["ymin"],
                x_max=b["xmax"],
                y_max=b["ymax"],
            ))
        return records

    # -- main --

    def run(self):
        label_map: Dict[str, List[BBoxRecord]] = {}
        dir_classes: Dict[str, List[str]] = {}
        for d in self._directories:
            dir_classes[d] = self._load_classes_for_dir(d)

        annotation_files: List[tuple] = []
        for d in self._directories:
            if not os.path.isdir(d):
                continue
            for root, _, files in os.walk(d):
                for fname in files:
                    ext = os.path.splitext(fname)[1].lower()
                    if ext == ".xml" or (ext == ".txt" and fname != "classes.txt"):
                        annotation_files.append((Path(root) / fname, d))

        size_cache: Dict[str, tuple] = {}
        total = len(annotation_files)
        processed = 0
        for fpath, parent_dir in annotation_files:
            if self._cancelled:
                break
            suffix = fpath.suffix.lower()
            records: List[BBoxRecord] = []
            if suffix == ".xml":
                records = self._parse_voc(fpath)
            elif suffix == ".txt":
                records = self._parse_yolo(
                    fpath, dir_classes.get(parent_dir, []), size_cache
                )

            for rec in records:
                label_map.setdefault(rec.label, []).append(rec)

            processed += 1
            if processed % 200 == 0 or processed == total:
                self.progress.emit(processed, total)

        self.finished.emit(label_map)


# ---------------------------------------------------------------------------
# GuideExportWorker
# ---------------------------------------------------------------------------

class GuideExportWorker(QThread):
    """按标签抽样裁剪、生成 Excel 标注指导文档。"""

    progress = pyqtSignal(int, int, str)  # current, total, status_text
    finished = pyqtSignal(bool, str)  # success, message/path

    def __init__(
        self,
        label_map: Dict[str, List[BBoxRecord]],
        selected_labels: List[str],
        output_path: str,
        samples_per_label: int = 5,
        expand_ratio: float = 2.0,
    ):
        super().__init__()
        self._label_map = label_map
        self._selected_labels = selected_labels
        self._output_path = output_path
        self._samples_per_label = samples_per_label
        self._expand_ratio = expand_ratio
        self._cancelled = False

    def cancel(self):
        self._cancelled = True

    @staticmethod
    def _crop_expanded(
        img: np.ndarray, bbox: BBoxRecord, expand_ratio: float
    ) -> Optional[np.ndarray]:
        """以 bbox 中心为基准扩大裁剪，并在截图上绘制 bbox 框线。"""
        h, w = img.shape[:2]
        bw = bbox.x_max - bbox.x_min
        bh = bbox.y_max - bbox.y_min
        if bw <= 0 or bh <= 0:
            return None
        cx = (bbox.x_min + bbox.x_max) / 2
        cy = (bbox.y_min + bbox.y_max) / 2
        new_w = bw * expand_ratio
        new_h = bh * expand_ratio
        x1 = max(0, int(cx - new_w / 2))
        y1 = max(0, int(cy - new_h / 2))
        x2 = min(w, int(cx + new_w / 2))
        y2 = min(h, int(cy + new_h / 2))
        if x2 <= x1 or y2 <= y1:
            return None
        crop = img[y1:y2, x1:x2].copy()

        bx1 = bbox.x_min - x1
        by1 = bbox.y_min - y1
        bx2 = bbox.x_max - x1
        by2 = bbox.y_max - y1
        thickness = max(1, min(crop.shape[:2]) // 150)
        cv2.rectangle(crop, (bx1, by1), (bx2, by2), (0, 0, 255), thickness)
        return crop

    def run(self):
        try:
            self._generate(self._output_path)
        except Exception as e:
            logger.exception("Guide export failed")
            self.finished.emit(False, str(e))

    @staticmethod
    def _encode_to_bytesio(img: np.ndarray) -> Optional[io.BytesIO]:
        """将 OpenCV 图像编码为 JPEG 并返回 BytesIO 对象。"""
        success, buf = cv2.imencode(".jpg", img, [cv2.IMWRITE_JPEG_QUALITY, 85])
        if not success:
            return None
        bio = io.BytesIO(buf.tobytes())
        bio.seek(0)
        return bio

    def _generate(self, output_path: str):
        wb = Workbook()
        wb.remove(wb.active)

        total_labels = len(self._selected_labels)
        done = 0

        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        wrap_align = Alignment(wrap_text=True, vertical="center", horizontal="center")

        img_cache: Dict[str, np.ndarray] = {}
        used_names: Set[str] = set()

        for label in self._selected_labels:
            if self._cancelled:
                self.finished.emit(False, "已取消")
                return

            records = self._label_map.get(label, [])
            sample_count = min(self._samples_per_label, len(records))
            sampled = random.sample(records, sample_count) if records else []

            safe_name = _SHEET_NAME_INVALID_RE.sub("_", label)[:28].strip()
            base = safe_name or "unknown"
            name = base
            counter = 1
            while name in used_names:
                suffix = f"_{counter}"
                name = base[:28 - len(suffix)] + suffix
                counter += 1
            used_names.add(name)
            ws = wb.create_sheet(title=name)

            # -- header row --
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(sample_count, 1))
            header_cell = ws.cell(row=1, column=1)
            header_cell.value = f"标签: {label}    总实例数: {len(records)}    抽样: {sample_count}"
            header_cell.font = header_font_white
            header_cell.fill = header_fill
            header_cell.alignment = wrap_align
            ws.row_dimensions[1].height = 30

            # -- sub-header with column indices --
            for col_idx in range(sample_count):
                cell = ws.cell(row=2, column=col_idx + 1)
                cell.value = f"示例 {col_idx + 1}"
                cell.font = header_font
                cell.alignment = wrap_align

            # -- image row --
            img_row = 3
            max_img_h = 0

            for col_idx, rec in enumerate(sampled):
                if self._cancelled:
                    self.finished.emit(False, "已取消")
                    return

                try:
                    raw = img_cache.pop(rec.image_path, None)
                    if raw is None:
                        raw = imread_unicode(rec.image_path)
                        if raw is None:
                            logger.warning("Cannot read image: %s", rec.image_path)
                            continue
                        if len(img_cache) >= _IMG_CACHE_MAX:
                            del img_cache[next(iter(img_cache))]
                    img_cache[rec.image_path] = raw

                    crop = self._crop_expanded(raw, rec, self._expand_ratio)
                    if crop is None:
                        continue

                    max_dim = 400
                    ch, cw = crop.shape[:2]
                    if max(ch, cw) > max_dim:
                        scale = max_dim / max(ch, cw)
                        crop = cv2.resize(crop, (int(cw * scale), int(ch * scale)))

                    img_bio = self._encode_to_bytesio(crop)
                    if img_bio is None:
                        continue

                    xl_img = XlImage(img_bio)
                    ch, cw = crop.shape[:2]
                    max_w, max_h = 300, 260
                    scale = min(max_w / cw, max_h / ch)
                    xl_img.width = int(cw * scale)
                    xl_img.height = int(ch * scale)
                    col_letter = get_column_letter(col_idx + 1)
                    ws.column_dimensions[col_letter].width = max(
                        ws.column_dimensions[col_letter].width or 0,
                        xl_img.width / 7 + 2,
                    )
                    ws.add_image(xl_img, f"{col_letter}{img_row}")
                    max_img_h = max(max_img_h, xl_img.height)
                except Exception:
                    logger.debug("Failed to process bbox for %s", rec.image_path, exc_info=True)

            ws.row_dimensions[img_row].height = max(max_img_h * 0.75 + 15, 100)

            # -- source info row --
            info_row = img_row + 1
            ws.row_dimensions[info_row].height = 50
            for col_idx, rec in enumerate(sampled):
                cell = ws.cell(row=info_row, column=col_idx + 1)
                cell.value = Path(rec.image_path).name
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
                cell.font = Font(size=9, color="666666")

            done += 1
            self.progress.emit(done, total_labels, f"正在处理: {label}")

        try:
            wb.save(output_path)
            self.finished.emit(True, output_path)
        except Exception as e:
            self.finished.emit(False, f"保存失败: {e}")


# ---------------------------------------------------------------------------
# AnnotationGuidePage
# ---------------------------------------------------------------------------

class AnnotationGuidePage(QWidget):
    """标注指导文档导出页面。"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._project_manager = None
        self._label_map: Dict[str, List[BBoxRecord]] = {}
        self._scan_worker: Optional[LabelScanWorker] = None
        self._export_worker: Optional[GuideExportWorker] = None
        self._archive_to_projects: Dict[str, Set[str]] = {}
        self._setup_ui()

    def set_project_manager(self, pm):
        self._project_manager = pm
        self._refresh_dataset_list()

    # ------------------------------------------------------------------
    # UI setup
    # ------------------------------------------------------------------

    def _setup_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)

        scroll_area = ScrollArea(self)
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("QScrollArea { border: none; background: transparent; }")

        content = QWidget(self)
        self._content_layout = QVBoxLayout(content)
        self._content_layout.setContentsMargins(36, 20, 36, 20)
        self._content_layout.setSpacing(16)

        self._content_layout.addWidget(TitleLabel("标注指导文档", self))
        self._content_layout.addWidget(self._create_dataset_card())
        self._content_layout.addWidget(self._create_label_card())
        self._content_layout.addWidget(self._create_config_card())
        self._content_layout.addWidget(self._create_export_card())
        self._content_layout.addStretch()

        scroll_area.setWidget(content)
        main_layout.addWidget(scroll_area)

    # -- cards --

    def _create_dataset_card(self) -> CardWidget:
        card = CardWidget(self)
        layout = QVBoxLayout(card)
        layout.setContentsMargins(20, 16, 20, 16)
        layout.setSpacing(10)

        layout.addWidget(SubtitleLabel("数据集选择", card))
        layout.addWidget(CaptionLabel("勾选一个或多个数据集项目（归档会自动展开为子项目）", card))

        btn_row = QHBoxLayout()
        self._refresh_btn = PushButton("刷新列表", card)
        self._refresh_btn.setIcon(FIF.SYNC)
        self._refresh_btn.clicked.connect(self._refresh_dataset_list)
        btn_row.addWidget(self._refresh_btn)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        self._dataset_list = QListWidget(card)
        self._dataset_list.setMinimumHeight(140)
        self._dataset_list.setMaximumHeight(260)
        self._dataset_list.itemChanged.connect(self._on_dataset_item_changed)
        layout.addWidget(self._dataset_list)

        return card

    def _create_label_card(self) -> CardWidget:
        card = CardWidget(self)
        layout = QVBoxLayout(card)
        layout.setContentsMargins(20, 16, 20, 16)
        layout.setSpacing(10)

        layout.addWidget(SubtitleLabel("标签选择", card))
        layout.addWidget(CaptionLabel("先扫描选中数据集的标注，然后勾选需要导出的标签", card))

        btn_row = QHBoxLayout()
        self._scan_btn = PrimaryPushButton("扫描标签", card)
        self._scan_btn.setIcon(FIF.SEARCH)
        self._scan_btn.clicked.connect(self._start_scan)
        btn_row.addWidget(self._scan_btn)

        self._select_all_btn = PushButton("全选", card)
        self._select_all_btn.clicked.connect(lambda: self._toggle_all_labels(True))
        btn_row.addWidget(self._select_all_btn)

        self._deselect_all_btn = PushButton("全不选", card)
        self._deselect_all_btn.clicked.connect(lambda: self._toggle_all_labels(False))
        btn_row.addWidget(self._deselect_all_btn)

        btn_row.addStretch()
        layout.addLayout(btn_row)

        self._scan_progress = ProgressBar(card)
        self._scan_progress.setVisible(False)
        layout.addWidget(self._scan_progress)

        self._scan_status = CaptionLabel("尚未扫描", card)
        layout.addWidget(self._scan_status)

        self._label_list = QListWidget(card)
        self._label_list.setMinimumHeight(120)
        self._label_list.setMaximumHeight(300)
        layout.addWidget(self._label_list)

        return card

    def _create_config_card(self) -> CardWidget:
        card = CardWidget(self)
        layout = QVBoxLayout(card)
        layout.setContentsMargins(20, 16, 20, 16)
        layout.setSpacing(10)

        layout.addWidget(SubtitleLabel("导出配置", card))

        row1 = QHBoxLayout()
        row1.addWidget(BodyLabel("每类标签抽样数量", card))
        self._samples_spin = SpinBox(card)
        self._samples_spin.setRange(1, 50)
        self._samples_spin.setValue(5)
        row1.addWidget(self._samples_spin)
        row1.addStretch()
        layout.addLayout(row1)

        row2 = QHBoxLayout()
        row2.addWidget(BodyLabel("截取扩展比例", card))
        self._expand_spin = DoubleSpinBox(card)
        self._expand_spin.setRange(1.5, 5.0)
        self._expand_spin.setValue(2.5)
        self._expand_spin.setSingleStep(0.5)
        row2.addWidget(self._expand_spin)
        row2.addStretch()
        layout.addLayout(row2)

        layout.addWidget(CaptionLabel(
            "扩展比例越大，截取区域越大，能看到更多目标周围的上下文环境", card
        ))

        return card

    def _create_export_card(self) -> CardWidget:
        card = CardWidget(self)
        layout = QVBoxLayout(card)
        layout.setContentsMargins(20, 16, 20, 16)
        layout.setSpacing(10)

        layout.addWidget(SubtitleLabel("导出", card))

        btn_row = QHBoxLayout()
        self._export_btn = PrimaryPushButton("导出 Excel", card)
        self._export_btn.setIcon(FIF.SAVE)
        self._export_btn.clicked.connect(self._start_export)
        btn_row.addWidget(self._export_btn)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        self._export_progress = ProgressBar(card)
        self._export_progress.setVisible(False)
        layout.addWidget(self._export_progress)

        self._export_status = CaptionLabel("", card)
        layout.addWidget(self._export_status)

        return card

    # ------------------------------------------------------------------
    # Dataset list
    # ------------------------------------------------------------------

    def _refresh_dataset_list(self):
        self._dataset_list.blockSignals(True)
        self._dataset_list.clear()
        self._archive_to_projects.clear()
        if self._project_manager is None:
            self._dataset_list.blockSignals(False)
            return

        pm = self._project_manager
        added_ids: Set[str] = set()

        for arc in pm.archives.values():
            item = QListWidgetItem(f"[归档] {arc.name}")
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Unchecked)
            item.setData(Qt.UserRole, ("archive", arc.id))
            self._dataset_list.addItem(item)
            added_ids.update(arc.project_ids)
            self._archive_to_projects[arc.id] = set(arc.project_ids)

        for proj in pm.projects.values():
            if proj.is_virtual or proj.id.startswith("archive::"):
                continue
            prefix = "  ↳ " if proj.id in added_ids else ""
            item = QListWidgetItem(f"{prefix}{proj.name}  ({proj.directory})")
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Unchecked)
            item.setData(Qt.UserRole, ("project", proj.id))
            self._dataset_list.addItem(item)
        self._dataset_list.blockSignals(False)

    def _on_dataset_item_changed(self, item: QListWidgetItem):
        kind, obj_id = item.data(Qt.UserRole)
        if kind != "archive":
            return
        child_ids = self._archive_to_projects.get(obj_id)
        if not child_ids:
            return
        state = item.checkState()
        self._dataset_list.blockSignals(True)
        for i in range(self._dataset_list.count()):
            child = self._dataset_list.item(i)
            child_kind, child_id = child.data(Qt.UserRole)
            if child_kind == "project" and child_id in child_ids:
                child.setCheckState(state)
        self._dataset_list.blockSignals(False)

    def _get_selected_directories(self) -> List[str]:
        """返回所有勾选项对应的目录列表（去重）。"""
        if self._project_manager is None:
            return []
        pm = self._project_manager
        dirs: List[str] = []
        seen = set()

        for i in range(self._dataset_list.count()):
            item = self._dataset_list.item(i)
            if item.checkState() != Qt.Checked:
                continue
            kind, obj_id = item.data(Qt.UserRole)
            if kind == "archive":
                arc = pm.archives.get(obj_id)
                if arc:
                    for pid in arc.project_ids:
                        proj = pm.projects.get(pid)
                        if proj and proj.directory not in seen:
                            dirs.append(proj.directory)
                            seen.add(proj.directory)
            else:
                proj = pm.projects.get(obj_id)
                if proj and proj.directory not in seen:
                    dirs.append(proj.directory)
                    seen.add(proj.directory)
        return dirs

    # ------------------------------------------------------------------
    # Label scanning
    # ------------------------------------------------------------------

    def _start_scan(self):
        dirs = self._get_selected_directories()
        if not dirs:
            InfoBar.warning("提示", "请先勾选至少一个数据集", parent=self,
                            position=InfoBarPosition.TOP, duration=3000)
            return

        if self._scan_worker and self._scan_worker.isRunning():
            self._scan_worker.cancel()
            self._scan_worker.wait(3000)

        self._label_list.clear()
        self._label_map.clear()
        self._scan_btn.setEnabled(False)
        self._scan_progress.setVisible(True)
        self._scan_progress.setValue(0)
        self._scan_status.setText("正在扫描标注文件…")

        self._scan_worker = LabelScanWorker(dirs)
        self._scan_worker.progress.connect(self._on_scan_progress)
        self._scan_worker.finished.connect(self._on_scan_finished)
        self._scan_worker.start()

    def _on_scan_progress(self, current: int, total: int):
        if total > 0:
            self._scan_progress.setValue(int(current / total * 100))
        self._scan_status.setText(f"已扫描 {current} / {total} 个文件")

    def _on_scan_finished(self, label_map: Dict[str, List[BBoxRecord]]):
        self._label_map = label_map
        self._scan_btn.setEnabled(True)
        self._scan_progress.setVisible(False)

        self._label_list.clear()
        total_boxes = sum(len(v) for v in label_map.values())
        self._scan_status.setText(
            f"扫描完成：发现 {len(label_map)} 种标签，共 {total_boxes} 个标注框"
        )

        for label in sorted(label_map.keys()):
            count = len(label_map[label])
            item = QListWidgetItem(f"{label}  ({count} 个)")
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Checked)
            item.setData(Qt.UserRole, label)
            self._label_list.addItem(item)

    def _toggle_all_labels(self, checked: bool):
        state = Qt.Checked if checked else Qt.Unchecked
        for i in range(self._label_list.count()):
            self._label_list.item(i).setCheckState(state)

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------

    def _get_selected_labels(self) -> List[str]:
        labels = []
        for i in range(self._label_list.count()):
            item = self._label_list.item(i)
            if item.checkState() == Qt.Checked:
                labels.append(item.data(Qt.UserRole))
        return labels

    def _start_export(self):
        selected = self._get_selected_labels()
        if not selected:
            InfoBar.warning("提示", "请先扫描标签并勾选至少一个标签", parent=self,
                            position=InfoBarPosition.TOP, duration=3000)
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "保存标注指导文档", "标注指导文档.xlsx",
            "Excel 文件 (*.xlsx)"
        )
        if not path:
            return

        if self._export_worker and self._export_worker.isRunning():
            self._export_worker.cancel()
            self._export_worker.wait(3000)

        self._export_btn.setEnabled(False)
        self._export_progress.setVisible(True)
        self._export_progress.setValue(0)
        self._export_status.setText("正在生成…")

        self._export_worker = GuideExportWorker(
            label_map=self._label_map,
            selected_labels=selected,
            output_path=path,
            samples_per_label=self._samples_spin.value(),
            expand_ratio=self._expand_spin.value(),
        )
        self._export_worker.progress.connect(self._on_export_progress)
        self._export_worker.finished.connect(self._on_export_finished)
        self._export_worker.start()

    def _on_export_progress(self, current: int, total: int, status: str):
        if total > 0:
            self._export_progress.setValue(int(current / total * 100))
        self._export_status.setText(status)

    def _on_export_finished(self, success: bool, message: str):
        self._export_btn.setEnabled(True)
        self._export_progress.setVisible(False)
        if success:
            self._export_status.setText(f"导出完成: {message}")
            InfoBar.success("成功", "标注指导文档已导出", parent=self,
                            position=InfoBarPosition.TOP, duration=5000)
        else:
            self._export_status.setText(f"导出失败: {message}")
            InfoBar.error("失败", message, parent=self,
                          position=InfoBarPosition.TOP, duration=5000)
